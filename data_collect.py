import numpy as np
import cv2
import math
import scipy
# import imutils
# import matplotlib.pyplot as plt
from itertools import accumulate
from typing import List
import csv
import openpyxl
import os              
import time            

        
def calculate(image_gray,rectangle_long,rectangle_wide):
    img_long=image_gray.shape[1]
    img_wide=image_gray.shape[0]
    arr=np.empty((img_wide,img_long),dtype=np.int16)
    
    arr=image_gray
    #還原深度
    arr=arr.astype('uint16')
    arr=arr/255*4000
    arr_center=arr[int(img_wide/2),int(img_long/2)]
    
    dx = arr_center
    degree=180/math.pi
    ###FOV無黑點
    a_o=math.atan(172/224)
    DFOV_o=98/2/degree
    dh_o=dx/math.cos(DFOV_o)
    dy_o=dh_o*math.sin(DFOV_o)
    
    cx_o=dy_o*math.cos(a_o)
    cy_o=dy_o*math.sin(a_o)
    
    ###FOV計算
    cx=rectangle_long*cx_o/224
    cy=rectangle_wide*cy_o/172
    VFOV=math.atan(cy/dx)*degree*2
    HFOV=math.atan(cx/dx)*degree*2
        
    return HFOV,VFOV,dx

def maximalRectangle( array: List[List[str]]) -> int:
    if len(array) == 0:
            return 0
    arr = [list(map(int, x)) for x in array]
    n = len(array[0])
    up, left, right = [0] * n, [0] * n, [0] * n
    ans = 0
    long_perfect=0
    wide_perfect=0
    for row in arr:
        row_left = list(accumulate(row, lambda val, x: (val + x) * x))
        row_right = list(accumulate(row[::-1], lambda val, x: (val + x) * x))[::-1]
        up = [(val + x) * x for val, x in zip(up, row)]
        left = [min(x, y) if u > 1 else y for x, y, u in zip(left, row_left, up)]
        right = [min(x, y) if u > 1 else y for x, y, u in zip(right, row_right, up)]
        for u, l, r in zip(up, left, right):
            if ans<u*(l+r-1):
                long_perfect=l+r-1
                wide_perfect=u
            ans = max(ans, u * (l + r - 1))

                
    return ans,wide_perfect,long_perfect


if __name__ == '__main__':

	#Excel setting
	# 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    Dis1 = workbook.create_sheet('range1_1.0m')
    Dis2 = workbook.create_sheet("range1_1.5m")
    Dis3 = workbook.create_sheet("range1_2.0m")
    Dis4 = workbook.create_sheet("range1_2.5m")
    Dis5 = workbook.create_sheet("range1_3.0m")
    Dis6 = workbook.create_sheet('range2_1.0m')
    Dis7 = workbook.create_sheet("range2_1.5m")
    Dis8 = workbook.create_sheet("range2_2.0m")
    Dis9 = workbook.create_sheet("range2_2.5m")
    Dis10 = workbook.create_sheet("range2_3.0m")
    
    Alpha =  ['A','B','C','D','E','F','G','H','I','VFOV','HFOV','亮度_A','亮度_B','亮度_C','亮度_D','亮度_E','亮度_F','亮度_G','亮度_H','亮度_I']
    FieldName = ['B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1']
    WorkBookName = [Dis1,Dis2,Dis3,Dis4,Dis5,Dis6,Dis7,Dis8,Dis9,Dis10]

    for i in range(10):
        for j in range(20):
            WorkBookName[i][FieldName[j]] = Alpha[j]  

    Mode = input("PLz Enter Mode.\n A:1m ; B:1.5m ; C:2m ; D:2.5m ; E:3m \n")
    print("The Mode you choose is : ",Mode)

    ######改range#####   
    if(Mode in ["A" , "B" , "C" , "D" , "E"]):
        tof_range=1
    elif(Mode in ["F" , "G" , "H" , "I" , "J"]):
        tof_range=2
        
    # range_set="voxel3d_tools.exe -R "+str(tof_range)
    # os.system("C:/Users/work9/Desktop/voxel3d_tools-main/platform/win/Bin/x64-Debug/voxel3d_tools/"+range_set)

	# OPEN ToF CAMERA
    countFrame = 2
    cap = cv2.VideoCapture(0)
    np.uint16
    cap.set(cv2.CAP_PROP_CONVERT_RGB , 0)
    cap.set(cv2.CAP_PROP_MODE , 3)

    width = 224 #224 / 80
    height = 172 #172 / 60

    imIR 	= np.zeros((height,width) , dtype = "uint8")
    imDepth = np.zeros((height,width) , dtype = "uint8")

    imDepth_Raw = np.zeros((height * 2 , width) , dtype = "uint8")
    imIR_Raw 	= np.zeros((height * 2 , width) , dtype = "uint8")
    '''
	START FUNCTION --------------------------------------------------
    '''
    record = 0
    output = []
    output_list = []
    point_size = 3
    point_color = (0, 0, 255) # BGR
    thickness = 4 # 可以為 0 、4、8
    cv2.namedWindow("imIR",cv2.WINDOW_FREERATIO)
    cv2.namedWindow("imDepth",cv2.WINDOW_FREERATIO)
    cv2.resizeWindow("imIR",240,180)
    cv2.resizeWindow("imDepth",240,180)

    num = 0                                #YL 2022 12 12 start
    localtime = time.localtime()
    path = './TOF_try/image/'+time.strftime("%Y_%m_%d_%I_%M",localtime)
    xlsx_path = time.strftime("%Y_%m_%d_%I_%M",localtime)
    if not os.path.isdir(path):
    	os.makedirs(path)                  #YL 2022 12 12 END
        
        

    for i in range(2):
        if (i==1):
            tof_range=2
            if Mode == "A":
                Mode="F"
            elif Mode == "B":
                Mode="G"
            elif Mode == "C":
                Mode="H"
            elif Mode == "D":
                Mode="I"
            elif Mode == "E":
                Mode="J"
        elif (i==0):
            tof_range=1
        # range_set="voxel3d_tools.exe -R "+str(tof_range)
        # os.system("C:/Users/work9/Desktop/voxel3d_tools-main/platform/win/Bin/x64-Debug/voxel3d_tools/"+range_set)
        
        countFrame = 2
        cap = cv2.VideoCapture(0)
        np.uint16
        cap.set(cv2.CAP_PROP_CONVERT_RGB , 0)
        cap.set(cv2.CAP_PROP_MODE , 3)

        width = 224 #224 / 80
        height = 172 #172 / 60

        imIR 	= np.zeros((height,width) , dtype = "uint8")
        imDepth = np.zeros((height,width) , dtype = "uint8")

        imDepth_Raw = np.zeros((height * 2 , width) , dtype = "uint8")
        imIR_Raw 	= np.zeros((height * 2 , width) , dtype = "uint8")
        '''
        START FUNCTION --------------------------------------------------
        '''
        record = 0
        output = []
        output_list = []
        point_size = 3
        point_color = (0, 0, 255) # BGR
        thickness = 4 # 可以為 0 、4、8
        cv2.namedWindow("imIR",cv2.WINDOW_FREERATIO)
        cv2.namedWindow("imDepth",cv2.WINDOW_FREERATIO)
        cv2.resizeWindow("imIR",240,180)
        cv2.resizeWindow("imDepth",240,180)
            
        
        while (1):
		# GET IMAGES
            ret,rawMat = cap.read()

            x=np.frombuffer(rawMat , dtype = 'uint16').reshape(height*2 , width)

            cv2.convertScaleAbs(x , imIR_Raw , 255.0/4000.0 , 0)
            cv2.convertScaleAbs(x , imDepth_Raw , 255.0/4000.0 , 0)

            imIR             = imIR_Raw[height:height*2,0:width]
            imDepth          = imDepth_Raw[0:height , 0:width]
            imDepth_bgr      = cv2.cvtColor(imDepth, cv2.COLOR_GRAY2BGR)
            imIR_bgr         = cv2.cvtColor(imIR, cv2.COLOR_GRAY2BGR)
            DepthValue 		 = x[0:height,0:width]
            IRValue          = x[171:343,0:width]
            index_x, index_y = int(height/2),int(width/2)
        
            ############################################################################
            img2 = imDepth.copy()
            kernel_binary = np.ones((4,4),np.float32)
            dst=cv2.filter2D(img2, -1, kernel_binary)
            ret, img2_binary=cv2.threshold(dst, 1, 255, cv2.THRESH_BINARY)
            kernel_erosion = np.ones((10,10),np.float32)
            erosion = cv2.erode(img2_binary, kernel_erosion, iterations = 1)
            erosion_black_edge = cv2.copyMakeBorder(erosion,5,5,5,5, cv2.BORDER_CONSTANT,value=(0,0,0))
		
            rectangle_para=maximalRectangle(erosion_black_edge/255)
            rectangle_long=rectangle_para[2]
            rectangle_wide=rectangle_para[1]
            
            
                
            FOV=calculate(img2,rectangle_long,rectangle_wide)
            if FOV[2] != 0:
                VFOV=FOV[1]
                VFOV=round(VFOV)
                HFOV=FOV[0]
                HFOV=round(HFOV)
                center=FOV[2]
            else:
                VFOV=0
                HFOV=0
            
            img_center_IR=imIR.shape
		
            brightness=imIR[int(img_center_IR[0]/2),int(img_center_IR[1]/2)]
            brightness=brightness.astype('uint16')
            brightness=round(brightness*4000/255)
            ############################################################################
            
		#*************************************************************************#

            pt_list 		= [(0,0),(112,0),(223,0),(56,43),(168,43),(0,86),(112,86),(223,86),(56,129),(168,129),(0,171),(112,171),(223,171)]
            JoshPoint 		= [(37,28),(112,28),(186,28),(37,86),(112,86),(186,86),(37,143),(112,143),(186,143)]       
            Wayne_pt_list 	= [(0,0),(39,0),(79,0),(19,14),(58,14),(0,29),(39,29),(79,29),(19,44),(58,44),(0,59),(39,59),(79,59)] #80X60

            if width == 80:
                DepthPoint = [(13,10),(40,10),(66,10),(13,30),(40,30),(66,30),(13,50),(40,50),(66,50)]
            else:
                DepthPoint = [(37,28),(112,28),(186,28),(37,86),(112,86),(186,86),(37,143),(112,143),(186,143)]	


            for point in DepthPoint:
                cv2.circle(imIR_bgr, point, point_size, point_color, 0)
                cv2.circle(imDepth_bgr, point, point_size, point_color, 0)
                '''
                StingTI13point =  "p1  = %-14s"%str(DepthValue[0][0])+"p2  = %-14s"%str(DepthValue[index_x][0])+"p3  = %-14s"%str(DepthValue[-1][0])+"\n"\
                +"p4  = %-14s"%str(DepthValue[19][14])+"p5  = %-14s"%str(DepthValue[58][14])+"\n"\
                +"p6  = %-14s"%str(DepthValue[0][index_y])+"p7  = %-14s"%str(DepthValue[index_x][index_y])+"p8  = %-14s"%str(DepthValue[-1][index_y])+"\n"\
                +"p9  = %-14s"%str(DepthValue[19][44])+"p10 = %-14s"%str(DepthValue[58][44])+"\n"\
                +"p11 = %-14s"%str(DepthValue[0][-1])+"p12 = %-14s"%str(DepthValue[-index_x][-1])+"p13 = %-14s"%str(DepthValue[-1][-1])+'\n'
                '''
                Waynetext = \
                    str(DepthValue[0][0])+' '+str(DepthValue[index_x][0])+' '+str(DepthValue[-1][0])+' '+str(DepthValue[19][14])+' '\
                    +str(DepthValue[58][14])+' '+str(DepthValue[0][index_y])+' '+str(DepthValue[index_x][index_y])+' '+str(DepthValue[-1][index_y])+' '\
                    +str(DepthValue[19][44])+' '+str(DepthValue[58][44])+' '+str(DepthValue[0][-1])+' '+str(DepthValue[index_x][-1])+' '+str(DepthValue[-1][-1])+'\n'



                if width == 80 :
                    StingTI =  "p1  = %-14s"%str(DepthValue[10][13])+"p2  = %-14s"%str(DepthValue[10][40])+"p3  = %-14s"%str(DepthValue[10][66])+"\n"\
                    +"p4  = %-14s"%str(DepthValue[30][13])+"p5  = %-14s"%str(DepthValue[30][40])+"p6  = %-14s"%str(DepthValue[30][66])+"\n"\
                        +"p7  = %-14s"%str(DepthValue[50][13])+"p8  = %-14s"%str(DepthValue[50][40])+"p9  = %-14s"%str(DepthValue[50][66])+"\n"
                    p1 = DepthValue[10][13] 
                    p2 = DepthValue[10][40] 
                    p3 = DepthValue[10][66] 
                    p4 = DepthValue[30][13] 
                    p5 = DepthValue[30][40] 
                    p6 = DepthValue[30][66] 
                    p7 = DepthValue[50][13] 
                    p8 = DepthValue[50][40] 
                    p9 = DepthValue[50][66]

                else:
                        StringL723 =  "p1  = %-14s"%str(DepthValue[28][37])+"p2  = %-14s"%str(DepthValue[28][112])+"p3  = %-14s"%str(DepthValue[28][186])+"\n"\
                                    +"p4  = %-14s"%str(DepthValue[86][37])+"p5  = %-14s"%str(DepthValue[86][112])+"p6  = %-14s"%str(DepthValue[86][186])+"\n"\
                                    +"p7 = %-14s"%str(DepthValue[143][37])+"p8 = %-14s"%str(DepthValue[143][112])+"p9 = %-14s"%str(DepthValue[143][186])+'\n'
                        p1 = DepthValue[28][37] 
                        p2 = DepthValue[28][112] 
                        p3 = DepthValue[28][186] 
                        p4 = DepthValue[86][37] 
                        p5 = DepthValue[86][112] 
                        p6 = DepthValue[86][186] 
                        p7 = DepthValue[143][37] 
                        p8 = DepthValue[143][112] 
                        p9 = DepthValue[143][186]
                        
            StringL723 =  "p1  = %-14s"%str(IRValue[28][37])+"p2  = %-14s"%str(IRValue[28][112])+"p3  = %-14s"%str(IRValue[28][186])+"\n"\
                        +"p4  = %-14s"%str(IRValue[86][37])+"p5  = %-14s"%str(IRValue[86][112])+"p6  = %-14s"%str(IRValue[86][186])+"\n"\
                        +"p7 = %-14s"%str(IRValue[143][37])+"p8 = %-14s"%str(IRValue[143][112])+"p9 = %-14s"%str(IRValue[143][186])+'\n'
            IR_p1 = IRValue[28][37] 
            IR_p2 = IRValue[28][112] 
            IR_p3 = IRValue[28][186] 
            IR_p4 = IRValue[86][37] 
            IR_p5 = IRValue[86][112] 
            IR_p6 = IRValue[86][186] 
            IR_p7 = IRValue[143][37] 
            IR_p8 = IRValue[143][112] 
            IR_p9 = IRValue[143][186]

                

            pointList = [p1,p2,p3,p4,p5,p6,p7,p8,p9,VFOV,HFOV,IR_p1,IR_p2,IR_p3,IR_p4,IR_p5,IR_p6,IR_p7,IR_p8,IR_p9]
            if record == 1:
                tt = []
                if width == 80 :
                    print("Hello")
                    output.append(StingTI) 
                else:
                    output.append(StringL723) 
                for i,j in Wayne_pt_list:
                    tt.append(DepthValue[j][i])
                output_list.append(tt)
			
            cv2.imshow("imIR",imIR_bgr)
            cv2.imshow("imDepth",imDepth_bgr)


            if Mode=="A":
                for FieldABC in range(20):
                    Dis1.cell(countFrame, FieldABC+2).value = pointList[FieldABC]
                    
            elif Mode=="B":
                for FieldABC in range(20):
                    Dis2.cell(countFrame, FieldABC+2).value = pointList[FieldABC]

            elif Mode=="C":
                for FieldABC in range(20):
                    Dis3.cell(countFrame, FieldABC+2).value = pointList[FieldABC]

            elif Mode=="D":
                for FieldABC in range(20):
                    Dis4.cell(countFrame, FieldABC+2).value = pointList[FieldABC]

            elif Mode=="E":
                for FieldABC in range(20):
                    Dis5.cell(countFrame, FieldABC+2).value = pointList[FieldABC]
                
            elif Mode=="F":
                    for FieldABC in range(20):
                        Dis6.cell(countFrame, FieldABC+2).value = pointList[FieldABC]
                
            elif Mode=="G":
                for FieldABC in range(20):
                    Dis7.cell(countFrame, FieldABC+2).value = pointList[FieldABC]
                
            elif Mode=="H":
                for FieldABC in range(20):
                    Dis8.cell(countFrame, FieldABC+2).value = pointList[FieldABC]
                
            elif Mode=="I":
                for FieldABC in range(20):
                    Dis9.cell(countFrame, FieldABC+2).value = pointList[FieldABC]
                
            elif Mode=="J":
                    for FieldABC in range(20):
                        Dis10.cell(countFrame, FieldABC+2).value = pointList[FieldABC]
                
            else:
                    print("Type Error")
                    break
            if countFrame == 9001:    ##儲存9000筆資料
                print(str(countFrame) + " times iteration over !")
                workbook.save(path+'/'+xlsx_path+'.xlsx')
                
                cv2.imwrite(path+'/IR'+str(num)+'.jpg',imIR_bgr)
                cv2.imwrite(path+'/Depth'+str(num)+'.jpg',imDepth_bgr)
                num+=1
                break

            countFrame = countFrame + 1
		
            if cv2.waitKey(10) & 0xFF==ord('s'):  ##儲存影像
            
                cv2.imwrite(path+'/IR'+str(num)+'.jpg',imIR_bgr)
                cv2.imwrite(path+'/Depth'+str(num)+'.jpg',imDepth_bgr)
                
                num+=1
                print("TRUE")
            
            if cv2.waitKey(1) & 0xFF==ord('q'):   ##離開迴圈
                break
            
        cv2.destroyAllWindows()
        cap.release()
